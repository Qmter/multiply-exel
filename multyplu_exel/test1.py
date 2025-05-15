import sys
import pandas as pd
import re
from openpyxl import load_workbook
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QPushButton, QVBoxLayout, QWidget,
    QFileDialog, QMessageBox, QLabel, QHBoxLayout, QInputDialog
)
from PyQt6.QtCore import Qt
from datetime import datetime

# Кортеж с названиями месяцев на русском с годом
month = (
    "ЯНВАРЬ 2025", "ФЕВРАЛЬ 2025", "МАРТ 2025", "АПРЕЛЬ 2025", "МАЙ 2025",
    "ИЮНЬ 2025", "ИЮЛЬ 2025", "АВГУСТ 2025", "СЕНТЯБРЬ 2025", "ОКТЯБРЬ 2025",
    "НОЯБРЬ 2025", "ДЕКАБРЬ 2025")

# Словарь для перевода сокращений месяцев с английского на русский
month_en_to_rus = {
    "JAN": "ЯНВАРЬ", "FEB": "ФЕВРАЛЬ", "MAR": "МАРТ", "APR": "АПРЕЛЬ",
    "MAY": "МАЙ", "JUN": "ИЮНЬ", "JUL": "ИЮЛЬ", "AUG": "АВГУСТ",
    "SEP": "СЕНТЯБРЬ", "OCT": "ОКТЯБРЬ", "NOV": "НОЯБРЬ", "DEC": "ДЕКАБРЬ"
}

# Словарь для перевода русских названий месяцев в их числовое представление
month_rus_to_num = {
    "ЯНВАРЬ": "01", "ФЕВРАЛЬ": "02", "МАРТ": "03", "АПРЕЛЬ": "04",
    "МАЙ": "05", "ИЮНЬ": "06", "ИЮЛЬ": "07", "АВГУСТ": "08",
    "СЕНТЯБРЬ": "09", "ОКТЯБРЬ": "10", "НОЯБРЬ": "11", "ДЕКАБРЬ": "12"
}

month_num_to_en = {
    "01": "JAN",
    "02": "FEB",
    "03": "MAR",
    "04": "APR",
    "05": "MAY",
    "06": "JUN",
    "07": "JUL",
    "08": "AUG",
    "09": "SEP",
    "10": "OCT",
    "11": "NOV",
    "12": "DEC"
}


class UniversalExcelUpdater(QMainWindow):
    """Главный класс приложения для обработки Excel файлов."""

    def __init__(self):
        """Инициализация главного окна приложения."""
        super().__init__()
        self.init_ui()  # Настройка интерфейса
        self.source_file = None  # Путь к исходному файлу
        self.target_file = None  # Путь к целевому файлу

    def init_ui(self):
        """Настройка пользовательского интерфейса."""
        self.setWindowTitle("Универсальный обработчик Excel")
        self.setMinimumSize(400, 300)
        self.setMaximumSize(400, 300)

        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)

        # Метки для отображения выбранных файлов
        self.lbl_source = QLabel("Файл EX1: не выбран")
        self.lbl_target = QLabel("Файл EX2: не выбран")
        self.lbl_source.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_target.setAlignment(Qt.AlignmentFlag.AlignCenter)

        # Кнопки интерфейса
        self.btn_load_source = QPushButton("Загрузить файл с ID (EX1)")
        self.btn_load_target = QPushButton("Загрузить целевой файл (EX2)")
        self.btn_process = QPushButton("Запустить обработку")

        # Установка фиксированного размера для кнопок
        for btn in [self.btn_load_source, self.btn_load_target, self.btn_process]:
            btn.setFixedSize(200, 80)

        # Горизонтальные layout'ы для центрирования кнопок
        h_layout1 = QHBoxLayout()
        h_layout2 = QHBoxLayout()
        h_layout3 = QHBoxLayout()

        # Добавление кнопок в layout'ы с растягивающими элементами
        h_layout1.addStretch()
        h_layout1.addWidget(self.btn_load_source)
        h_layout1.addStretch()

        h_layout2.addStretch()
        h_layout2.addWidget(self.btn_load_target)
        h_layout2.addStretch()

        h_layout3.addStretch()
        h_layout3.addWidget(self.btn_process)
        h_layout3.addStretch()

        # Добавление всех элементов в основной layout
        main_layout.addWidget(self.lbl_source)
        main_layout.addLayout(h_layout1)
        main_layout.addWidget(self.lbl_target)
        main_layout.addLayout(h_layout2)
        main_layout.addLayout(h_layout3)
        main_layout.addStretch()

        # Подключение сигналов кнопок
        self.btn_load_source.clicked.connect(lambda: self.load_file('source'))
        self.btn_load_target.clicked.connect(lambda: self.load_file('target'))
        self.btn_process.clicked.connect(self.process_data)

    def load_file(self, file_type):
        """Загрузка файла через диалоговое окно."""
        file_name, _ = QFileDialog.getOpenFileName(
            self,
            "Выберите файл",
            "",
            "Excel Files (*.xlsx *.xls)"
        )

        if file_name:
            if file_type == 'source':
                self.source_file = file_name
                self.lbl_source.setText(f"EX1: {file_name.split('/')[-1]}")
            else:
                self.target_file = file_name
                self.lbl_target.setText(f"EX2: {file_name.split('/')[-1]}")

    def process_data(self):
        """Основной метод обработки данных."""
        try:
            # Чтение исходного файла в DataFrame
            self.df_source = pd.read_excel(self.source_file, header=None, dtype=object, engine='openpyxl')
            day_with_id_mapping = {}

            # Получение номера колонки для обработки
            column_source = self.letter_to_number(self.choose_column_source_dialog())

            column_target = self.choose_column_target_dialog()

            column_target_id = self.choose_column_target_id_dialog()


            # Чтение целевого файла
            self.df_target = pd.read_excel(self.target_file, header=None, dtype=object, engine='openpyxl')

            # Открытие файлов с помощью openpyxl
            wb1 = load_workbook(self.source_file)
            wb2 = load_workbook(self.target_file)

            # Получение названий листов через диалоговые окна
            source_sheet = self.choose_source_sheet_dialog(wb1.sheetnames)
            target_sheet = self.choose_target_sheet_dialog(wb2.sheetnames)

            # Заполнение словаря сопоставлений
            for i in source_sheet:
                id_mapping = {}  # Словарь для сопоставления паспортных данных с ID
                for _, row in pd.read_excel(self.source_file, header=None, dtype=object, engine='openpyxl',
                                            sheet_name=i).iterrows():
                    passport = str(row[column_source - 1]).strip().replace('.0', '') if not pd.isna(row[column_source - 1]) else ''
                    if passport:
                        id_mapping[passport] = row[0]
                        day_with_id_mapping[i] = id_mapping

            print("__________________________________________________")
            print(f"ID сопоставлений c днями: {day_with_id_mapping}")

            # Проверка соответствия листов
            self.sheet_check(source_sheet, target_sheet)

            print("____________________________________________________")
            self.add_id(target_sheet, day_with_id_mapping, column_target, column_target_id)

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка обработки:\n{str(e)}")

    def is_header_row_openpyxl(self, row):
        """Проверка, является ли строка заголовком."""
        try:
            first_cell = row[0].value
            if isinstance(first_cell, datetime):
                return all(cell.value is None for cell in row[1:])
            return False
        except:
            return False

    def letter_to_number(self, column_letter):
        """Конвертация буквенного обозначения колонки в числовое."""
        num = 0
        for i, char in enumerate(reversed(column_letter.upper())):
            num += (ord(char) - ord('A') + 1) * (26 ** i)
        return num

    def choose_column_source_dialog(self):
        """Диалог для выбора колонки."""
        text, ok = QInputDialog.getText(self, "Ввод названия колонки сравнения исходного файла", "Введите колонку для сравнения из исходного файла(A-ZZZ):")
        if ok:
            return text

    def choose_column_target_dialog(self):
        """Диалог для выбора колонки."""
        text, ok = QInputDialog.getInt(self, "Ввод названия колонки сравнения изменяемого файла", "Введите колонку для сравнения для изменяемого файла(A-ZZZ):")
        if ok:
            return text

    def choose_column_target_id_dialog(self):
        """Диалог для выбора колонки."""
        text, ok = QInputDialog.getText(self, "Ввод названия колонки (куда вставить id)", "Введите колонку для сравнения для вставки id)(A-ZZZ):")
        if ok:
            return text

    def choose_source_sheet_dialog(self, items):
        """Диалог для выбора листа с ID."""
        nes_date = []
        item, ok = QInputDialog.getText(self, f"Выбор из листа с ID", f"Выберите листы {items}:")
        if ok:
            for i in item.split(","):
                nes_date.append(i.strip())
            return nes_date

    def choose_target_sheet_dialog(self, items):
        """Диалог для выбора целевого листа."""
        item, ok = QInputDialog.getItem(self, "Выбор из списка", "Выберите вариант:", items, 0, False)
        if ok:
            return item

    def sheet_check(self, sheet1, sheet2):
        print("sheet_check")
        """Проверка соответствия листов."""
        try:
            for i in sheet1:
                date = []
                sheet1_day, sheet1_month = i.split()
                month_in_russian, _ = sheet2.split()

                if any(month_in_russian in s for s in month):
                    # Сбор дат из целевого файла

                    for j, row in pd.read_excel(self.target_file, header=None, dtype=object, engine='openpyxl',
                                                sheet_name=sheet2).iterrows():
                        if pd.isna(row[1]):
                            date.append(str(row[0]).split()[0])

                    self.date_check(sheet1_day, month_in_russian, self.form_date(date))
                else:
                    print("Месяц не найден в списке.")
        except ValueError:
            print("Ошибка формата ввода. Пример: '01 DEC'")

    def date_check(self, sheet1_day, sheet1_month, date):
        """Проверка соответствия дат."""
        print(date)
        global month_num, day
        for i in date:
            day, month_num, year = i.split(".")
            month_num_from_rus = month_rus_to_num.get(sheet1_month, "00")

            if sheet1_day == day and month_num == month_num_from_rus:
                print("ВСЁ ВЕРНО")
            else:
                print("ОШИБКА: Даты нет в листе")

    def form_date(self, dates):
        """Форматирование дат."""
        date_ = []
        for i in dates:
            date_obj = datetime.strptime(i, "%Y-%m-%d")
            formatted_date = date_obj.strftime("%d.%m.%Y")
            date_.append(formatted_date)
        return date_

    def form_date_add_id(self, dates):
        """Форматирование дат."""
        date_obj = datetime.strptime(str(dates), "%Y-%m-%d")
        formatted_date = date_obj.strftime("%d.%m.%Y")
        return formatted_date

    def add_id(self, list_to, id_mapping, pas_col_sheet2, id_col_sheet2):
        print(f"{list_to}, \n{id_mapping}, \n{pas_col_sheet2}, \n{id_col_sheet2}")
        wb = load_workbook(self.target_file.split('/')[-1])
        print(self.target_file.split('/')[-1])



        ws = wb[list_to]
        date = id_mapping.keys() if len(id_mapping.keys()) == 1 else ''

        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if pd.isna(row[1]):
                day, month, _ = self.form_date_add_id(row[0].date()).split('.')
                en_m = month_num_to_en[month]
                tmp = f"{day} {en_m}"
                if tmp in id_mapping.keys():
                    date = f"{day} {en_m}"
            if str(row[pas_col_sheet2]) in id_mapping[date].keys():
                ws[f'{id_col_sheet2}{i}'] = id_mapping[date][str(row[pas_col_sheet2])]

        wb.save('EX2.xlsx')

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = UniversalExcelUpdater()
    window.show()
    sys.exit(app.exec())
