import sys
import pandas as pd
import re
from openpyxl import load_workbook
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QPushButton, QVBoxLayout, QWidget,
    QFileDialog, QMessageBox, QLabel, QHBoxLayout
)
from PyQt6.QtCore import Qt
from datetime import datetime


class UniversalExcelUpdater(QMainWindow):
    def __init__(self):
        super().__init__()
        self.init_ui()
        self.source_file = None
        self.target_file = None

    def init_ui(self):
        self.setWindowTitle("Универсальный обработчик Excel")
        self.setMinimumSize(400, 300)
        self.setMaximumSize(400, 300)

        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)

        # Метки (по центру)
        self.lbl_source = QLabel("Файл EX1: не выбран")
        self.lbl_target = QLabel("Файл EX2: не выбран")
        self.lbl_source.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_target.setAlignment(Qt.AlignmentFlag.AlignCenter)

        # Создаем кнопки с фиксированным размером
        self.btn_load_source = QPushButton("Загрузить файл с ID (EX1)")
        self.btn_load_target = QPushButton("Загрузить целевой файл (EX2)")
        self.btn_process = QPushButton("Запустить обработку")

        # Устанавливаем фиксированный размер для всех кнопок
        for btn in [self.btn_load_source, self.btn_load_target, self.btn_process]:
            btn.setFixedSize(200, 80)

        # Создаем горизонтальные контейнеры для центрирования кнопок
        h_layout1 = QHBoxLayout()
        h_layout2 = QHBoxLayout()
        h_layout3 = QHBoxLayout()

        # Добавляем кнопки в горизонтальные layout'ы с растягиванием по бокам
        h_layout1.addStretch()
        h_layout1.addWidget(self.btn_load_source)
        h_layout1.addStretch()

        h_layout2.addStretch()
        h_layout2.addWidget(self.btn_load_target)
        h_layout2.addStretch()

        h_layout3.addStretch()
        h_layout3.addWidget(self.btn_process)
        h_layout3.addStretch()

        # Добавляем все элементы в основной вертикальный layout
        main_layout.addWidget(self.lbl_source)
        main_layout.addLayout(h_layout1)
        main_layout.addWidget(self.lbl_target)
        main_layout.addLayout(h_layout2)
        main_layout.addLayout(h_layout3)
        main_layout.addStretch()

        # Сигналы
        self.btn_load_source.clicked.connect(lambda: self.load_file('source'))
        self.btn_load_target.clicked.connect(lambda: self.load_file('target'))
        self.btn_process.clicked.connect(self.process_data)

    def load_file(self, file_type):
        file_name, _ = QFileDialog.getOpenFileName(
            self,
            "Выберите файл",
            "",
            "Excel Files (*.xlsx *.xls)"
        )


        if file_name:
            if file_type == 'source':
                self.source_file = file_name
                self.lbl_source.setText(f"EX1: {file_name.split('/')[-1]}")
            else:
                self.target_file = file_name
                self.lbl_target.setText(f"EX2: {file_name.split('/')[-1]}")

    def process_data(self):
        try:
            # Чтение исходного файла
            df_source = pd.read_excel(self.source_file, header=None, dtype=object, engine='openpyxl')
            id_mapping = {}

            # Создание словаря сопоставлений паспортных данных с ID
            for _, row in df_source.iterrows():
                passport = str(row[2]).strip().replace('.0', '') if not pd.isna(row[2]) else ''
                if passport:
                    id_mapping[passport] = row[0]

            print(f"ID сопоставлений из исходного файла: {id_mapping}")  # Выводим сопоставления для отладки

            # Открытие целевого файла
            wb = load_workbook(self.target_file)
            ws = wb.active

            added_count = 0

            # Проходим по строкам целевого файла
            for row_idx, row in enumerate(ws.iter_rows(min_row=1), start=1):
                if self.is_header_row_openpyxl(row):
                    continue

                matched_id = None

                # Находим индекс первого не пустого столбца в строке
                first_non_empty_column_idx = None
                for col_idx, cell in enumerate(row):
                    if cell.value not in [None, '']:  # Если ячейка не пуста
                        first_non_empty_column_idx = col_idx
                        break

                # Если нашли первый не пустой столбец, пропускаем его
                if first_non_empty_column_idx is not None:
                    print(f"Пропускаем столбец {first_non_empty_column_idx + 1} (индексация с 1) в строке {row_idx}")

                # Ищем паспортные данные в строках, пропуская первый непустой столбец
                for col_idx, cell in enumerate(row):
                    if col_idx == first_non_empty_column_idx:  # Пропускаем первый непустой столбец
                        continue

                    if cell.value is not None:
                        passport = str(cell.value).strip().replace('.0', '')
                        print(f"Сравниваю паспорт: {passport}")  # Выводим паспорта для отладки
                        if passport in id_mapping:
                            matched_id = id_mapping[passport]
                            break

                # Если нашли соответствие, вставляем ID
                if matched_id is not None:
                    print(f"Найдено совпадение для паспорта: {passport}, ID: {matched_id}")  # Для отладки

                    # Проверяем, есть ли уже ID в строке (извлекая первое число перед фамилией)
                    id_found_in_row = False
                    for col_idx, cell in enumerate(row):
                        if col_idx == first_non_empty_column_idx:  # Пропускаем первый непустой столбец
                            continue

                        if cell.value is not None:
                            # Преобразуем значение в строку и убираем лишние пробелы
                            existing_value = str(cell.value).strip().replace('.0', '')

                            # Используем регулярное выражение для извлечения первого числа в строке
                            match = re.match(r'^\d+', existing_value)  # Ищем только цифры в начале строки
                            if match:
                                existing_id = match.group(0)  # Получаем найденное число
                                print(
                                    f"Сравниваю с извлеченным значением: {existing_id} с {matched_id}")  # Выводим для отладки
                                if existing_id == str(matched_id):
                                    id_found_in_row = True
                                    break

                    if id_found_in_row:
                        continue  # Пропускаем, если ID уже есть в строке

                    # Вставляем ID в первый столбец после последнего непустого
                    last_col_idx = 1
                    for cell in row:
                        if cell.value is not None:
                            last_col_idx = cell.column
                    insert_col = last_col_idx + 1
                    ws.cell(row=row_idx, column=insert_col, value=matched_id)
                    added_count += 1

            # Сохраняем изменения
            wb.save(self.target_file)

            # Информируем пользователя
            QMessageBox.information(
                self,
                "Успех",
                f"Обработка завершена!\n"
                f"Добавлено ID: {added_count}"
            )

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка обработки:\n{str(e)}")

    def is_header_row_openpyxl(self, row):
        try:
            first_cell = row[0].value
            if isinstance(first_cell, datetime):
                return all(cell.value is None for cell in row[1:])
            return False
        except:
            return False



if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = UniversalExcelUpdater()
    window.show()
    sys.exit(app.exec())